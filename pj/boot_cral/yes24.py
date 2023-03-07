import requests as req
from bs4 import BeautifulSoup
import openpyxl

book_title=input("책 제목을 입력하시오 > ")
book_author=input("저자를 입력하시오 > ")

def yes24(book_title, book_author):
    try:
        data = req.get(f"http://www.yes24.com/product/search?domain=all&query={book_title}&page=1&_searchtarget=&_rqlist={book_author}")
        soup = BeautifulSoup(data.text, "html.parser")

        target_tittle=soup.select_one(".gd_name").text      #0
        target_author=soup.select_one(".authPub").text      #1
        target_author=target_author.strip() #좌우 공백제거
        target_price=soup.select_one(".yes_b").text+"원 ("+soup.select_one(".txt_sale").text+" 할인)"   #2
        return ["yes24",target_tittle, target_author, target_price]
    except:
        return ["yes24","잘못 입력하셨습니다", "", ""]

"""초기화
wb=openpyxl.Workbook()
ws = wb.active
ws.title = 'First_Sheet'
ws.cell(1, 1, value="No")
ws.cell(1, 2, value="store")
ws.cell(1, 3, value="tittle")
ws.cell(1, 4, value="author")
ws.cell(1, 5, value="price")
"""

wb = openpyxl.load_workbook("BookSearch.xlsx")     #execl 파일 이름
ws = wb.active                                      #작업할 워크시트
yes24_info=yes24(book_title, book_author)
k=ws.max_row
for i in range(len(yes24_info)):
    if i==0:
        ws.cell(k + 1, i+1, value=k)
        ws.cell(k + 1, i + 2, value=yes24_info[i])
    else:
        ws.cell(k+1, i+2, value=yes24_info[i])

wb.save("BookSearch.xlsx")