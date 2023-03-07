import requests as req
from bs4 import BeautifulSoup as bs4
from selenium import webdriver
from selenium.webdriver.common.by import By
import time
from urllib import parse
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service
import openpyxl

options = webdriver.ChromeOptions()
options.add_argument('headless')
#options.add_argument('user-agent: Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/108.0.0.0 Whale/3.18.154.13 Safari/537.36')
options.add_argument('window-size=1920x1080')
options.add_argument("disable-gpu")
service = Service(ChromeDriverManager().install())
driver = webdriver.Chrome(service=service, options=options)

#options.add_experimental_option("excludeSwitches", ["enable-logging"])

def yes24(book_title, book_author):
    try:
        data = req.get(f"http://www.yes24.com/product/search?domain=all&query={book_title}&page=1&_searchtarget=&_rqlist={book_author}")
        soup = bs4(data.text, "html.parser")

        target_title=soup.select_one(".gd_name").text      #1
        target_author=soup.select_one(".authPub > a").text      #2
        target_author=target_author.strip() #좌우 공백제거
        target_price=soup.select_one(".yes_b").text+"원 ("+soup.select_one(".txt_sale").text+" 할인)"   #3
        return ["yes24",target_title, target_author, target_price]
    except:
        return ["yes24","잘못 입력하셨습니다", "", ""]

def amazon(book_title, book_author):
    query_txt = book_title + " " + book_author
    driver.get("https://www.amazon.com/books-used-books-textbooks/b?ie=UTF8&node=283155")
    time.sleep(2)
    
    try:
        # 검색창 클릭
        driver.find_element(By.ID, "twotabsearchtextbox").click()
        element = driver.find_element(By.ID, "twotabsearchtextbox")

        # 검색어 입력
        element.send_keys(query_txt)

        # 검색
        driver.find_element(By.ID, "nav-search-submit-button").click()  
            
        # 도서 선택 후 가격 출력
        driver.find_element(By.CLASS_NAME, "s-image").click()
        price_titles = driver.find_element(By.CSS_SELECTOR, "#tmmSwatches > ul > li.swatchElement.selected")
        price = price_titles.find_element(By.CLASS_NAME, "a-color-base").text
        # price = price[5:]

        # 책 제목 출력
        titles = driver.find_element(By.CSS_SELECTOR, "#titleblock_feature_div > div")
        title = titles.find_element(By.ID, "title").text
        
        # 책 저자 출력
        authors = driver.find_element(By.CSS_SELECTOR, "#bylineInfo > span")
        author = authors.find_element(By.CLASS_NAME, "a-link-normal").text
        return ["amazon", title, author, price]
    except:
        return ["amazon", "잘못 입력하셨습니다", "", ""]
    driver.close()

def aladin(book_title, book_author):
    try:
        driver.get("https://www.aladin.co.kr/home/welcome.aspx")
        time.sleep(2)

        #driver.maximize_window # 창을 최대 크기로 키우기
        
        driver.find_element(By.ID, 'SearchWord').click()
        search = driver.find_element(By.ID, 'SearchWord')

        search.send_keys(book_title, book_author)
        time.sleep(1)
        # search.send_keys("\n")
        driver.find_element(By.CLASS_NAME, 'searchBtn').click()
        time.sleep(3)
        dis = driver.find_element(By.CLASS_NAME, 'ss_p').text
        driver.find_element(By.CLASS_NAME, 'bo3').click()
        titles = driver.find_element(By.CSS_SELECTOR, 'div > ul > li:nth-child(2) > div > a.Ere_bo_title').text
        prices = driver.find_element(By.CSS_SELECTOR, 'ul > li:nth-child(2) > div.Ritem.Ere_ht11 > span').text
        author = driver.find_element(By.CSS_SELECTOR, 'div > ul > li.Ere_sub2_title > a:nth-child(1)').text
        time.sleep(3)

        return ['알라딘', titles, author, prices+"원 ("+dis+" 할인)"]
    except:
        return ['알라딘', '잘못 입력하셨습니다', '', '']

def Kyobo(book_title, book_author):
    try:
        title_encoded = parse.quote(book_title)
        author_encoded = parse.quote(book_author)

        url = f'https://search.kyobobook.co.kr/search?keyword={title_encoded}&gbCode=KBO&target=kyobo&rekeyword={author_encoded}'
        res = req.get(url)
        html = bs4(res.text, 'html.parser')

        prod_info_box = html.select('.prod_info_box')
        prod_title = prod_info_box[0].select_one('a span[id*="cmdtName"]')
        prod_author = prod_info_box[0].select_one('div.prod_author_info a.author')
        if prod_author is None:
            prod_author = prod_info_box[0].select_one('div.prod_author_info span.author')
        prod_price = prod_info_box[0].select_one('span.val')
        normal_price = prod_info_box[0].select_one('s.val')
        discount = int((1 - float(prod_price.text.replace(',', '').split('원')[0]) / float(normal_price.text.replace(',', '').split('원')[0])) * 100.5)

        return ["교보문고", prod_title.text, prod_author.text, prod_price.text+"원 ("+str(discount)+"% 할인)"]
    except:
        return ["교보문고","잘못 입력하셨습니다", "", ""]

def interpark(book_title, book_author):
    try:
        driver.get('https://book.interpark.com')
        time.sleep(2)
        el = driver.find_element(By.ID,'query')
        el.send_keys(book_title, " ", book_author)
        el = driver.find_element(By.CLASS_NAME, "btn").click()

        soup = bs4(driver.page_source, "lxml")
        allcontents = soup.select('.list_wrap')
        if allcontents:
            for content in allcontents:
                title = content.select_one('b').text
                author = content.select_one('.writer > a').text
                price = content.select_one('.nowMoney').text

                title = title.replace("\n", "").replace("\r", "").replace("\t", "")
                author = author.replace("\n", "").replace("\r", "").replace("\t", "")
                price = price.replace("▼", " 할인").replace("(", " (")

                return["인터파크", title, author, price]
        else:
            return ["인터파크","잘못 입력하셨습니다", "", ""]
    except:
        return["인터파크","잘못 입력하셨습니다", "", ""]

def run():
    book_title=input("책 제목을 입력하시오 > ")
    book_author=input("저자를 입력하시오 > ")

    yes24_info=yes24(book_title, book_author)
    aladin_info = aladin(book_title, book_author)
    kyobo_info=Kyobo(book_title, book_author)
    interpark_info = interpark(book_title, book_author)
    amazon_info = amazon(book_title, book_author)

    print(yes24_info)
    print(aladin_info)
    print(kyobo_info)
    print(interpark_info)
    print(amazon_info)

    """초기화
    wb=openpyxl.Workbook()
    ws = wb.active
    ws.title = 'First_Sheet'
    ws.cell(1, 1, value="No")
    ws.cell(1, 2, value="store")
    ws.cell(1, 3, value="tittle")
    ws.cell(1, 4, value="author")
    ws.cell(1, 5, value="price")
    """
    wb = openpyxl.load_workbook("BookSearch.xlsx")  # execl 파일 이름
    ws = wb.active                                  # 작업할 워크시트

    k = ws.max_row
    #각 서점별로 함수로 만들어 wb, ws 열고 저장하는것을 반복하는것보다 더 낫지 않을까 생각
    for i in range(len(yes24_info)):
        if i == 0:
            if k < 12:
                ws.cell(k + 1, i + 1, value=k%4)
                ws.cell(k + 1, i + 2, value=yes24_info[i])
            else:
                ws.cell(k + 1, i + 1, value=k//4)
                ws.cell(k + 1, i + 2, value=yes24_info[i])
        else:
            ws.cell(k + 1, i + 2, value=yes24_info[i])
    k += 1
    for i in range(len(aladin_info)):
        if i == 0:
            ws.cell(k + 1, i + 2, value=aladin_info[i])
        else:
            ws.cell(k + 1, i + 2, value=aladin_info[i])
    k += 1
    for i in range(len(kyobo_info)):
        if i == 0:
            ws.cell(k + 1, i + 2, value=kyobo_info[i])
        else:
            ws.cell(k + 1, i + 2, value=kyobo_info[i])
    k += 1
    for i in range(len(interpark_info)):
        if i == 0:
            ws.cell(k + 1, i + 2, value=interpark_info[i])
        else:
            ws.cell(k + 1, i + 2, value=interpark_info[i])
    k += 1
    for i in range(len(amazon_info)):
        if i == 0:
            ws.cell(k + 1, i + 2, value=amazon_info[i])
        else:
            ws.cell(k + 1, i + 2, value=amazon_info[i])

    wb.save("BookSearch.xlsx")

run()
"""
book_title="ssssssssssssssssssssssssssssss"
book_author="kkkkkkkkkkkkkkkkkkkkkk"
interpark_info = interpark(book_title, book_author)
print(interpark_info)
"""